"""
ONTAP REST API Sample Scripts

Purpose: Script to list volumes properties using ONTAP REST API.

Usage: csa.py [-h] -s SITE_CODE -env PROD [-u API_USER] [-p API_PASS]
"""

import pandas as pd
import openpyxl as xl
import urllib3 as ur
import socket
import base64
import argparse
from getpass import getpass
from time import sleep
import logging
import texttable as tt
import requests
import json

import sys

ur.disable_warnings()

def find_clstr(site: str, envir: str, domain: str):
    """Get cluster info from inventory using user inputs"""
    
   #wb = xl.load_workbook(r'C:\\Users\\Administrator.DEMO\\Documents\\GitHub\\ProjA\\projclstrs.xlsx')
    wb = xl.load_workbook(r'/opt/storage_scripts/test_env_automation_v2/test_env_automation_v2/projclstrs.xlsx')
#active worksheet data
    ws = wb.active    
    

    output = []
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if site in ws.cell(i,j).value:
                val = ws.cell(i,j).value   
                if envir in val:
                    val = ws.cell(i,j).value
                    if domain == 'amz':
                       # list of amz tags
                       amz = ['bkp','cdoc','cf0','devi','erp0','nps','vm0']
                       for dom in amz:
                            if dom in val:
                                op = ''.join(val)
                                output.append(op)    
                    elif domain in val:
                        op = ''.join(val)
                        output.append(op)
    print(output)
    return output
    
def list_aggregate(cluster: str, dsktype: str, headers_inc: str) -> None:
    """Lists the Aggregate"""
    print()
    #print("List of Aggregates on ",cluster)
    #print("==========================================")
    r=0
    tab = tt.Texttable()
    header = ['Cluster Name','VServer Name','Aggr name','Size(GB)','Available(GB)','Used %']
    tab.header(header)
    tab.set_cols_width([20,25,35,15,15,10])
    tab.set_cols_align(['c','c','c','c','c','c'])

    for dsk in dsktype:
        url = "https://{}/api/storage/aggregates?name=*{}*".format(cluster,dsk)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
            tmp = dict(response.json())
            if "error" in tmp:
                print("Invalid Username/Password")
                print(tmp)
                sys.exit(1)
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
        tmp = dict(response.json())
        aggr = tmp['records']

        for i in aggr:
            r = r + 1
            aggr_uuid = i['uuid']
            url = "https://{}/api/storage/aggregates/{}".format(cluster,aggr_uuid)
            try:
                response = requests.get(url, headers=headers_inc, verify=False)
            except requests.exceptions.HTTPError as err:
                print(err)
                sys.exit(1)
            except requests.exceptions.RequestException as err:
                print(err)
                sys.exit(1)
            tmp = dict(response.json())
            tmp2 = dict(tmp['space'])
            tmp3 = dict(tmp2['block_storage'])
            avail = (((int(tmp3['available'])/1024)/1024)/1024)
            size = (((int(tmp3['size'])/1024)/1024)/1024)
            used = (((int(tmp3['used'])/1024)/1024)/1024)
            uip = (used * 100)/avail
            #uip = int(uip)
            aggr_name = i['name']
            svm_name = list_svm(cluster, headers)
            #print("svm_name is ", svm_name)
            tab.add_row([cluster,svm_name,aggr_name,size,avail,uip])
            tab.set_cols_width([20,25,35,15,15,10])
            tab.set_cols_align(['c','c','c','c','c','c'])
        #print("Number of Storage VMs on this NetApp cluster :{}".format(ctr))
    setdisplay = tab.draw()
    print(setdisplay)
        
def sort_svm(cluster: str, headers_inc: str):
    """Sorts VServers with app condition"""
    apps = ARGS.app
    services = ARGS.proto
    app_list = ["svm","arch","bkp","cdp","cf","dap","ddb","dmz","dp","dpt","erp","hd","mist","nps","pap","pdb","rdb","san","sris","tap","tdb","test","vm","cdoc","devi","sdb"]
    ctr = 0
    sort_row = tmp_n = []
    tmp = dict(get_vservers(cluster, headers_inc))
    vservers = tmp['records']
    
    for i in vservers:
        #print(i)
        ctr = ctr + 1
        if i is False:
            print("Vserver can't be sorted for app "+apps+", select one of different app value", app_list)
            sys.exit(1)
        if services == 'nfs':
            sort_name = i['name']
            tmp_n.append(sort_name)
            #print("fn sort_svm sort_name", sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
                    #print("fn sort_svm - sort_row", sort_row) 
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
            
        elif services == 'cifs':
            rcd_dt = dict(i)
            svm_rd = rcd_dt['svm']
            svm_dt = dict(svm_rd)
            sort_name = svm_dt['name']
            tmp_n.append(sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
        else:
            
            tmp=dict(i)
            tmp1 = tmp['svm']
            tmp3 = dict(tmp1)
            sort_name = tmp3['name']
            tmp_n.append(sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
            #sort_name = i['name']
    tmp_n = set(tmp_n)
    tmp_n = list(tmp_n)
    #print("tmp_n",tmp_n)
    sort_row = tmp_n
    
    #print("tmp list", tmp_n)
    #print("sort row value",sort_row)
        
            
    return sort_row 

    
def list_svm(cluster: str, headers_inc: str):
    """Lists the VServers"""
    hostname = ARGS.host 
    services = ARGS.proto
    apps = ARGS.app
    try:
        host_ip_add = socket.gethostbyname(hostname).split('.')
        host_subnet = '.'.join(host_ip_add[0:3])
    except socket.gaierror as err:
        host_subnet = "127.0.0"
        
                
    ctr = 0
    tmp = dict(get_vservers(cluster, headers_inc))
    #print(" tmp of list_svm ", tmp, ctr)
    vservers = tmp['records']
    srt = clus = []
    row = []
    for i in vservers:
        ctr = ctr + 1
        if services == 'nfs':
            clus = i['name']
            print("nfs clus val", clus)
            #print("fn list_svm - clus",clus)
            try:
                svm_ip_add = socket.gethostbyname(clus).split('.')
                svm_subnet = '.'.join(svm_ip_add[0:3])
            except socket.gaierror as err:
                svm_subnet = "0.0.0"
            
            if host_subnet == svm_subnet:
                row = [clus+"*"]
                return row
            srt = sort_svm(cluster, headers_inc)
            #print("srt output",srt)
            #clus = list(set(clus) | set(srt))
            #row = clus
            row = srt
        elif services == 'cifs':
            rcd_dt = dict(i)
            svm_rd = rcd_dt['svm']
            svm_dt = dict(svm_rd)
            clus = svm_dt['name']
            print("cifs clus val", clus)
            #clus = list(clus)
            srt = sort_svm(cluster, headers_inc)
            #print(srt)
            #print(clus)
            #clus = list(set(clus) | set(srt))
            row=srt
        elif services == 'iscsi':
            tmp=dict(i)
            tmp1 = tmp['svm']
            tmp3 = dict(tmp1)
            clus = tmp3['name']
            srt = sort_svm(cluster, headers_inc)
            #clus = list(set(clus) | set(srt))
            row=srt
        else:
            print("Enter valid protocol")
            sys.exit(1)
    
    tmp12 = set(row)
    tmp12 = list(tmp12)
    tmp12.sort()
    tmp10 = []
    for j in tmp12:
        if "afsx" in j:
            tmp12.remove(j)
        #elif "cf" in j:
         #   tmp10.append(j)
    #tmp10.sort()
    #print("tmp10 fin val", tmp10)
    print("tmp12",tmp12)
    
    for i in tmp12:
        if apps in i:
            tmp10.append(i)
        elif "cf" in i:
            tmp10.append(i)

    for chk in tmp10:
        adc = auth_dp_chk(cluster,chk,headers_inc)
        #print(adc)

        row_dt = dict(adc)
        chk_ind = row_dt['num_records']

        if chk_ind == 0:
            tmp10.remove(chk)

    sleep(3)      
    row = []
    #print("row",row)
    #
    #   
    for k in tmp10:
        print(" k ",k)
        
        if apps in k:
            row = []
            row.append(k)
            return row
        elif "cf" in k:
            #row = []
            row.append(k)
        #elif "afsx" in k and "afsx" in row:
        #    row.remove(k)
    #    elif "-dr" in k and "-dr" in row:
    #        row.remove(k)
    #    else:
    #        row.append(k)
            
    print("last finl row", row)        
    
   # for chk in row:
    #    adc = auth_dp_chk(cluster,chk,headers_inc)
        #print(adc)
        
     #   row_dt = dict(adc)
      #  chk_ind = row_dt['num_records']
       # 
        #if chk_ind == 0:
         #   row.remove(chk)
            
    #print("finl row", row)
    #sleep(3)
    print("retun row",row)
    return row


def auth_dp_chk(cluster: str, fsvm: str, headers_inc: str):
    """ excludes auth and dp destination SVM's """
    
    url = "https://{}/api/svm/svms?subtype=!dp_destination&name={}&return_timeout=15".format(cluster,fsvm)
    try:
        response = requests.get(url, headers=headers_inc, verify=False)
        #print(response.json())
    except requests.exceptions.HTTPError as err:
        print(err)
        sys.exit(1)
    except requests.exceptions.RequestException as err:
        print(err)
        sys.exit(1)
        
    return response.json()    
    
def get_vservers(cluster: str, headers_inc: str):
    """ Get vServer"""
    services = ARGS.proto
    
    if services == 'nfs':
        url = "https://{}/api/svm/svms?{}.enabled=true".format(cluster,services)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
            #print(response.json())
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    elif services == 'cifs':
        url = "https://{}/api/protocols/{}/services?enabled=true".format(cluster,services)
        
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
               
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    elif services == 'iscsi':
        url = "https://{}/api/protocols/san/{}/services?enabled=true".format(cluster,services)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    else:
        print()
        print(" Enter Valid protocol, should be nfs, cifs or iscsi")
        sys.exit(1)
        
        
    out_response = response.json() 
    
    return out_response
    
def parse_args() -> argparse.Namespace:
    """Parse the command line arguments from the user"""

    parser = argparse.ArgumentParser(
        description="This script will list volumes in a SVM")
    parser.add_argument(
        "-s", required=True, help="It should be valide site code like uslv,demu,usto so on.."
                        )
    parser.add_argument(
        "-env", required=True, help="It should be prod or nprod"  
                        )
    parser.add_argument(
        "-host", required=True, help="Valid Servername"  
                        )
    parser.add_argument(
        "-app", required=True, help="App name or purpose of provisioning like arch,bkp,cdp,cf0,pdb,ddb,dmz so on.."  
                        )                    
    parser.add_argument(
        "-proto", required=True, help="Valid protocal value of nfs,cifs,iscsi,fc"  
                        )    
    parser.add_argument(
        "-domain", required=True, help="Valid domain value of amz or dmz"  
                        )
    parser.add_argument(
        "-dskt", required=False, help="It should be sas,ssd or sata"  
                        )         
    parser.add_argument(
        "-u",
        "--api_user",
        default="admin",
        help="API Username")
    parser.add_argument("-p", "--api_pass", help="API Password")
    parsed_args = parser.parse_args()

    # collect the password without echo if not already provided
    if not parsed_args.api_pass:
        parsed_args.api_pass = getpass()

    return parsed_args

#def get_volumes(cluster: str, svm_name: str, volume_name: str, headers_inc: str):
#    """Get Volumes"""
#    url = "https://{}/api/storage/volumes/?svm.name={}".format(cluster, volume_name)
#    response = requests.get(url, headers=headers_inc, verify=False)
#    return response.json()

                
                
if __name__ == "__main__":

    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] [%(levelname)5s] [%(module)s:%(lineno)s] %(message)s",
    )
    ARGS = parse_args()
    BASE_64_STRING = base64.encodebytes(
        ('%s:%s' %
         (ARGS.api_user, ARGS.api_pass)).encode()).decode().replace('\n', '')
    
    headers = {
        'authorization': "Basic %s" % BASE_64_STRING,
        'content-type': "application/json",
        'accept': "application/json"
    }
    
    
    
    if (ARGS.env == 'prod' or ARGS.domain == 'dmz'):
        #ARGS.env = 'pfsx'
        if ARGS.dskt == 'sata':
            dsktype = ['sas','ssd','sata']
        else:
            dsktype = ['sas','ssd']
            if ARGS.env == 'nprod':
                ARGS.env = 'sfsx'
            else:
                ARGS.env = 'pfsx'
        #ARGS.env = 'pfsx'
        print(" if prod n dmz ", dsktype, ARGS.env)
        clstr_name = find_clstr(ARGS.s, ARGS.env, ARGS.domain)
        for clstr in clstr_name:
                aggr_list = list_aggregate(clstr,dsktype,headers)
                #svm_list = list_svm(clstr, headers)
    elif ARGS.env == 'nprod':
        ARGS.env = 'sfsx'
        if (ARGS.dskt == 'sas' or ARGS.dskt == 'ssd'):
            dsktype = ['sas','ssd','sata']
        else:
            dsktype = ['sata']
        #ARGS.env = 'sfsx'
        print(" if nprod n dmz or amz ", dsktype, ARGS.env)
        clstr_name = find_clstr(ARGS.s, ARGS.env, ARGS.domain)
        for clstr in clstr_name:
                aggr_list = list_aggregate(clstr,dsktype,headers)
                #svm_list = list_svm(clstr, headers)
    else:
        print()
        print("-env value invalid, it should be prod or nprod")
        sys.exit(1)
        

    
    
    
    #disp_vservers(ARGS.cluster, headers)
    
