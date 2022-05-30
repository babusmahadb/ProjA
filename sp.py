"""
ONTAP REST API Scripts

Purpose: Validation Script to list Cluster/VServer/Aggrerate details using ONTAP REST API.

Usage: sp.py [-h] -s SITE_NAME -env ENV -host HOST -app APP -proto PROTO -domain DOMAIN [-dskt DSKT] [-u API_USER] [-p API_PASS]
"""

import pandas as pd
import openpyxl as xl
import urllib3 as ur
import socket
import base64
import argparse
from getpass import getpass
from time import sleep
import logging
import texttable as tt
import requests
import json

import sys

ur.disable_warnings()

def find_clstr(site: str, envir: str, domain: str):
    """Get cluster info from inventory using user inputs"""
    
    wb = xl.load_workbook(r'C:\\Users\\Administrator.DEMO\\Documents\\GitHub\\ProjA\\projclstrs.xlsx')
    #wb = xl.load_workbook(r'/opt/storage_scripts/test_env_automation_v2/test_env_automation_v2/projclstrs.xlsx')
#active worksheet data
    ws = wb.active    
    

    output = []
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if site in ws.cell(i,j).value:
                val = ws.cell(i,j).value   
                if envir in val:
                    val = ws.cell(i,j).value
                    if domain == 'amz':
                       # list of amz tags
                       amz = ['bkp','cdoc','cf0','devi','erp0','nps','vm0']
                       for dom in amz:
                            if dom in val:
                                op = ''.join(val)
                                output.append(op)    
                    elif domain in val:
                        op = ''.join(val)
                        output.append(op)
    #print(output)
    return output
    
def list_aggregate(cluster: str, dsktype: str, headers_inc: str) -> None:
    """Lists the Aggregate"""
    print()
    #print("List of Aggregates on ",cluster)
    #print("==========================================")
    r=0
    tab = tt.Texttable()
    header = ['Cluster Name','VServer Name','Aggr name','Size(GB)','Available(GB)','Used %']
    tab.header(header)
    tab.set_cols_width([20,25,35,15,15,10])
    tab.set_cols_align(['c','c','c','c','c','c'])

    for dsk in dsktype:
        url = "https://{}/api/storage/aggregates?name=*{}*".format(cluster,dsk)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
            tmp = dict(response.json())
            if "error" in tmp:
                print("Invalid Username/Password")
                print(tmp)
                sys.exit(1)
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
        tmp = dict(response.json())
        aggr = tmp['records']

        for i in aggr:
            #print("Val of i", i)
            r = r + 1
            aggr_uuid = i['uuid']
            url = "https://{}/api/storage/aggregates/{}".format(cluster,aggr_uuid)
            try:
                response = requests.get(url, headers=headers_inc, verify=False)
            except requests.exceptions.HTTPError as err:
                print(err)
                sys.exit(1)
            except requests.exceptions.RequestException as err:
                print(err)
                sys.exit(1)
            tmp = dict(response.json())
            tmp2 = dict(tmp['space'])
            tmp3 = dict(tmp2['block_storage'])
            avail = (((int(tmp3['available'])/1024)/1024)/1024)
            size = (((int(tmp3['size'])/1024)/1024)/1024)
            used = (((int(tmp3['used'])/1024)/1024)/1024)
            uip = (used * 100)/size
            #uip = int(uip)
            aggr_name = i['name']
            svm_name = list_svm(cluster, headers)
            #print("svm_name is ", svm_name)
            tab.add_row([cluster,svm_name,aggr_name,size,avail,uip])
            tab.set_cols_width([20,25,35,15,15,10])
            tab.set_cols_align(['c','c','c','c','c','c'])
        #print("Number of Storage VMs on this NetApp cluster :{}".format(ctr))
    setdisplay = tab.draw()
    print(setdisplay)
    
    return svm_name
        
def sort_svm(cluster: str, headers_inc: str):
    """Sorts VServers with app condition"""
    apps = ARGS.app
    services = ARGS.proto
    app_list = ["svm","arch","bkp","cdp","cf","dap","ddb","dmz","dp","dpt","erp","hd","mist","nps","pap","pdb","rdb","san","sris","tap","tdb","test","vm","cdoc","devi","sdb"]
    ctr = 0
    sort_row = tmp_n = []
    tmp = dict(get_vservers(cluster, headers_inc))
    vservers = tmp['records']
    
    for i in vservers:
        #print(i)
        ctr = ctr + 1
        if i is False:
            print("Vserver can't be sorted for app "+apps+", select one of different app value", app_list)
            sys.exit(1)
        if services == 'nfs':
            sort_name = i['name']
            tmp_n.append(sort_name)
            #print("fn sort_svm sort_name", sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
                    #print("fn sort_svm - sort_row", sort_row) 
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
            
        elif services == 'cifs':
            rcd_dt = dict(i)
            svm_rd = rcd_dt['svm']
            svm_dt = dict(svm_rd)
            sort_name = svm_dt['name']
            tmp_n.append(sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
        else:
            
            tmp=dict(i)
            tmp1 = tmp['svm']
            tmp3 = dict(tmp1)
            sort_name = tmp3['name']
            tmp_n.append(sort_name)
            if apps in app_list:
                if apps in sort_name:
                    sort_row.append(sort_name)
            else:
                print("provide valid -app value, it must be one of ",app_list)
                sys.exit(1)
            #sort_name = i['name']
    tmp_n = set(tmp_n)
    tmp_n = list(tmp_n)
    #print("tmp_n",tmp_n)
    sort_row = tmp_n
    
    #print("tmp list", tmp_n)
    #print("sort row value",sort_row)
        
            
    return sort_row 

    
def list_svm(cluster: str, headers_inc: str):
    """Lists the VServers"""
    hostname = ARGS.host 
    services = ARGS.proto
    apps = ARGS.app
    try:
        host_ip_add = socket.gethostbyname(hostname).split('.')
        host_subnet = '.'.join(host_ip_add[0:3])
    except socket.gaierror as err:
        host_subnet = "127.0.0"
        
                
    ctr = 0
    tmp = dict(get_vservers(cluster, headers_inc))
    #print(" tmp of list_svm ", tmp, ctr)
    vservers = tmp['records']
    srt = clus = []
    row = []
    for i in vservers:
        ctr = ctr + 1
        if services == 'nfs':
            clus = i['name']
            #print("fn list_svm - clus",clus)
            try:
                svm_ip_add = socket.gethostbyname(clus).split('.')
                svm_subnet = '.'.join(svm_ip_add[0:3])
            except socket.gaierror as err:
                svm_subnet = "0.0.0"
            
            if host_subnet == svm_subnet:
                row = [clus+"*"]
                return row
            srt = sort_svm(cluster, headers_inc)
            #print("srt output",srt)
            #clus = list(set(clus) | set(srt))
            #row = clus
            row = srt
        elif services == 'cifs':
            rcd_dt = dict(i)
            svm_rd = rcd_dt['svm']
            svm_dt = dict(svm_rd)
            clus = svm_dt['name']
            #clus = list(clus)
            srt = sort_svm(cluster, headers_inc)
            #print(srt)
            #print(clus)
            #clus = list(set(clus) | set(srt))
            row=srt
        elif services == 'iscsi':
            tmp=dict(i)
            tmp1 = tmp['svm']
            tmp3 = dict(tmp1)
            clus = tmp3['name']
            srt = sort_svm(cluster, headers_inc)
            #clus = list(set(clus) | set(srt))
            row=srt
        else:
            print("Enter valid protocol")
            sys.exit(1)
    
    flt_vs = set(row)
    flt_vs = list(flt_vs)
    flt_vs.sort()
    upd_vs = []
    for j in flt_vs:
        if "afsx" in j:
            flt_vs.remove(j)
    for i in flt_vs:
        if apps in i:
            upd_vs.append(i)
        elif "cf" in i:
            upd_vs.append(i)

    for chk in upd_vs:
        adc = auth_dp_chk(cluster,chk,headers_inc)
        #print(adc)

        row_dt = dict(adc)
        chk_ind = row_dt['num_records']

        if chk_ind == 0:
            upd_vs.remove(chk)

    sleep(3)      
    row = []
      
    for k in upd_vs:
                
        if apps in k:
            row = []
            row.append(k)
            return row
        elif "cf" in k:
            row.append(k)
    
    #if ARGS.sm == 'y':
    #    
    #    row = []
    #    row = svm_peer
    #    print("snapm row", row)
    
    
    return row

def svm_peer(cluster: str, peer_cluster: str, headers_inc: str):

    svmp_url = "https://{}/api/svm/peers/?peer.cluster.name={}".format(cluster,peer_cluster)
    response = requests.get(svmp_url, headers=headers_inc, verify=False)
    svmp_json = response.json()
    
    svm_pr_dt = dict(svmp_json)
    svm_pr_rd = svm_pr_dt['records']
    
    svm_lst = []
    
    for r in svm_pr_rd:
        svm_pr_lt = r['name']
        svm_lst.append(svm_pr_lt)
        
    #print("peer svm", svm_lst)
    
    return svm_lst

def auth_dp_chk(cluster: str, fsvm: str, headers_inc: str):
    """ excludes auth and dp destination SVM's """
    
    url = "https://{}/api/svm/svms?subtype=!dp_destination&name={}&return_timeout=15".format(cluster,fsvm)
    try:
        response = requests.get(url, headers=headers_inc, verify=False)
        #print(response.json())
    except requests.exceptions.HTTPError as err:
        print(err)
        sys.exit(1)
    except requests.exceptions.RequestException as err:
        print(err)
        sys.exit(1)
        
    return response.json()    
    
def get_vservers(cluster: str, headers_inc: str):
    """ Get vServer"""
    services = ARGS.proto
    
    if services == 'nfs':
        url = "https://{}/api/svm/svms?{}.enabled=true".format(cluster,services)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
            #print(response.json())
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    elif services == 'cifs':
        url = "https://{}/api/protocols/{}/services?enabled=true".format(cluster,services)
        
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
               
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    elif services == 'iscsi':
        url = "https://{}/api/protocols/san/{}/services?enabled=true".format(cluster,services)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
    else:
        print()
        print(" Enter Valid protocol, should be nfs, cifs or iscsi")
        sys.exit(1)
        
        
    out_response = response.json() 
    
    return out_response


def get_clus_peer(cluster: str, headers_inc: str):
    """ get cluster peer details """

    cls_pr_url = "https://{}/api/cluster/peers?return_records=true&return_timeout=15".format(cluster)
    try:
        response = requests.get(cls_pr_url, headers=headers_inc, verify=False)
        #print(response.json())
    except requests.exceptions.HTTPError as err:
        print(err)
        sys.exit(1)
    except requests.exceptions.RequestException as err:
        print(err)
        sys.exit(1)
    
    cls_pr_json = response.json()
    
    cls_pr_dt = dict(cls_pr_json)
    cls_pr_rd = cls_pr_dt['records']
    
    cls_lst = []
    
    for r in cls_pr_rd:
        cls_pr_lt = r['name']
        cls_lst.append(cls_pr_lt)
        
    print("Peer Cluster for "+str(cluster)+" is "+str(cls_lst)+".")
    print()
    
    return cls_lst
        
def parse_args() -> argparse.Namespace:
    """Parse the command line arguments from the user"""

    parser = argparse.ArgumentParser(
        description="This script will list volumes in a SVM")
    parser.add_argument(
        "-s", required=True, help="It should be valide site code like uslv,demu,usto so on.."
                        )
    parser.add_argument(
        "-env", required=True, help="It should be prod or nprod"  
                        )
    parser.add_argument(
        "-host", required=True, help="Valid Servername"  
                        )
    parser.add_argument(
        "-app", required=True, help="App name or purpose of provisioning like arch,bkp,cdp,cf0,pdb,ddb,dmz so on.."  
                        )                    
    parser.add_argument(
        "-proto", required=True, help="Valid protocal value of nfs,cifs,iscsi,fc"  
                        )    
    parser.add_argument(
        "-domain", required=True, help="Valid domain value of amz or dmz"  
                        )
    parser.add_argument(
        "-dskt", required=False, help="It should be sas,ssd or sata"  
                        )
    parser.add_argument(
        "-sm", required=False, help="Pull Peer Cluster/SVM/Aggr Info"  
                        )                        
    parser.add_argument(
        "-u",
        "--api_user",
        default="admin",
        help="API Username")
    parser.add_argument("-p", "--api_pass", help="API Password")
    parsed_args = parser.parse_args()

    # collect the password without echo if not already provided
    if not parsed_args.api_pass:
        parsed_args.api_pass = getpass()

    return parsed_args

def snpchk(clstr: str, headers: str):
        
    print()
    clus_peer = get_clus_peer(clstr, headers)
    
    #print("Cluster/SVM Peer", clus_peer, svm_peer)
    
    while True:
        
        peer_clus = input("Enter a Peer Cluster name/IP for SnapMirror Configuration[Can be - uspa-pfsx-cf01, usas-pfsx-nps01]: ")
        
        if peer_clus in clus_peer:
            dsktype = ['sas','ssd','sata']
            peer_aggr_list = list_aggregate(peer_clus,dsktype,headers)
            svm_p = svm_peer(clstr, peer_clus, headers)
            break
        else:
            print()
            print("Entered Cluster is not peered with Source Cluster, Try these Cluster name: ", clus_peer)
            continue
    print()
    print("Source Cluster/SVM "+str(clstr)+"/"+str(aggr_list)+" Peered with "+str(peer_clus)+"/"+str(svm_p)+".")
    print()

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    
if __name__ == "__main__":

    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] [%(levelname)5s] [%(module)s:%(lineno)s] %(message)s",
    )
    ARGS = parse_args()
    BASE_64_STRING = base64.encodebytes(
        ('%s:%s' %
         (ARGS.api_user, ARGS.api_pass)).encode()).decode().replace('\n', '')
    
    headers = {
        'authorization': "Basic %s" % BASE_64_STRING,
        'content-type': "application/json",
        'accept': "application/json"
    }
    
    smirror = ARGS.sm
    
    if (ARGS.env == 'prod' or ARGS.domain == 'dmz'):
        if ARGS.dskt == 'sata':
            dsktype = ['sas','ssd','sata']
        else:
            dsktype = ['sas','ssd']
            if ARGS.env == 'nprod':
                ARGS.env = 'sfsx'
            else:
                ARGS.env = 'pfsx'
        
        clstr_name = find_clstr(ARGS.s, ARGS.env, ARGS.domain)
        for clstr in clstr_name:
                aggr_list = list_aggregate(clstr,dsktype,headers)
                if smirror == "y":
                    snpchk(clstr, headers)
                #svm_list = list_svm(clstr, headers)
    elif ARGS.env == 'nprod':
        if (ARGS.dskt == 'sas' or ARGS.dskt == 'ssd'):
            dsktype = ['sas','ssd','sata']
        else:
            dsktype = ['sata']
        ARGS.env = 'sfsx'
        clstr_name = find_clstr(ARGS.s, ARGS.env, ARGS.domain)
        for clstr in clstr_name:
                aggr_list = list_aggregate(clstr,dsktype,headers)
                if smirror == "y":
                    snpchk(clstr, headers)
                #svm_list = list_svm(clstr, headers)
    else:
        print()
        print("-env value invalid, it should be prod or nprod")
        sys.exit(1)

    #clus_peer = get_clus_peer(clstr, headers)
    




