"""
ONTAP REST API Sample Scripts

Purpose: Script to list volumes properties using ONTAP REST API.

Usage: csa.py [-h] -s SITE_CODE -env PROD [-u API_USER] [-p API_PASS]
"""

import pandas as pd
import openpyxl as xl
import urllib3 as ur

import base64
import argparse
from getpass import getpass
import logging
#import texttable as tt
import requests

import sys

ur.disable_warnings()

def find_clstr(site: str, envir: str):
    """Get cluster info from inventory using user inputs"""
    
    wb = xl.load_workbook(r'C:\\Users\\Administrator.DEMO\\Documents\\GitHub\\test\\amgenclstrs.xlsx')

#active worksheet data
    ws = wb.active    

    output = []
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if site in ws.cell(i,j).value:
                #print("found")
                val = ws.cell(i,j).value   
                if envir in val:
                   op = ''.join(val)
                   output.append(op)

    return output
    

def list_aggregate(cluster: str, headers_inc: str) -> None:
    """Lists the Aggregate"""
    url = "https://{}/api/storage/aggregates".format(cluster)
    try:
        response = requests.get(url, headers=headers_inc, verify=False)
    except requests.exceptions.HTTPError as err:
        print(err)
        sys.exit(1)
    except requests.exceptions.RequestException as err:
        print(err)
        sys.exit(1)
    tmp = dict(response.json())
    aggr = tmp['records']
    #bs = tmp['records.block_storage']
    #print()
    #print(ns)
    print()
    print("List of Aggregates on ",cluster)
    print("=====================")
    for i in aggr:
        aggr_uuid = i['uuid']
        url = "https://{}/api/storage/aggregates/{}".format(cluster,aggr_uuid)
        try:
            response = requests.get(url, headers=headers_inc, verify=False)
        except requests.exceptions.HTTPError as err:
            print(err)
            sys.exit(1)
        except requests.exceptions.RequestException as err:
            print(err)
            sys.exit(1)
        tmp = dict(response.json())
        tmp2 = dict(tmp['space'])
        tmp3 = dict(tmp2['block_storage'])
        avail = tmp3['available']
        space_convert=(((int(avail)/1024)/1024)/1024)
        #aggr = tmp['records']
        print("Aggregate %s " % i['name'] +"has available space of "+str(int(space_convert))+" GB")
        #print("Aggregate UUID = %s " % i['uuid'])
        
        
def parse_args() -> argparse.Namespace:
    """Parse the command line arguments from the user"""

    parser = argparse.ArgumentParser(
        description="This script will list volumes in a SVM")
    parser.add_argument(
        "-s", "--site", required=True, help="It should be valide site code like uslv,demu,usto so on.."
                        )
    parser.add_argument(
        "-env", "--envir", required=True, help="It should be pfsx or sfsx"  
                        )
    parser.add_argument(
        "-u",
        "--api_user",
        default="admin",
        help="API Username")
    parser.add_argument("-p", "--api_pass", help="API Password")
    parsed_args = parser.parse_args()

    # collect the password without echo if not already provided
    if not parsed_args.api_pass:
        parsed_args.api_pass = getpass()

    return parsed_args

#def get_volumes(cluster: str, svm_name: str, volume_name: str, headers_inc: str):
#    """Get Volumes"""
#    url = "https://{}/api/storage/volumes/?svm.name={}".format(cluster, volume_name)
#    response = requests.get(url, headers=headers_inc, verify=False)
#    return response.json()

                
                
if __name__ == "__main__":

    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] [%(levelname)5s] [%(module)s:%(lineno)s] %(message)s",
    )
    ARGS = parse_args()
    BASE_64_STRING = base64.encodebytes(
        ('%s:%s' %
         (ARGS.api_user, ARGS.api_pass)).encode()).decode().replace('\n', '')
    
    headers = {
        'authorization': "Basic %s" % BASE_64_STRING,
        'content-type': "application/json",
        'accept': "application/json"
    }
    
    clstr_name = find_clstr(ARGS.site, ARGS.envir)
    
    aggr_list = list_aggregate(clstr_name[0],headers)
    